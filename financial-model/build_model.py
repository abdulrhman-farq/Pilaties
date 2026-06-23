# -*- coding: utf-8 -*-
"""
بناء النموذج المالي التفاعلي لاستوديو Reformer Pilates النسائي - الرياض
يُولّد ملف Excel بصيغ مترابطة: تعديل خلايا الافتراضات يُحدّث كل المخرجات تلقائياً.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== الأنماط =====
TITLE = Font(name="Arial", size=14, bold=True, color="FFFFFF")
HEAD = Font(name="Arial", size=11, bold=True, color="FFFFFF")
BOLD = Font(name="Arial", size=11, bold=True)
NORM = Font(name="Arial", size=11)
INPUT_FONT = Font(name="Arial", size=11, bold=True, color="1F4E78")

NAVY = PatternFill("solid", fgColor="1F4E78")
TEAL = PatternFill("solid", fgColor="2E8B8B")
LGREY = PatternFill("solid", fgColor="D9E1F2")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

wb = openpyxl.Workbook()
wb.properties.title = "النموذج المالي - استوديو ريفورمر بيلاتس الرياض"


def style_row(ws, row, c1, c2, fill=None, font=NORM, border=True, align=None):
    for col in range(c1, c2 + 1):
        cell = ws.cell(row=row, column=col)
        if fill:
            cell.fill = fill
        cell.font = font
        if border:
            cell.border = BORDER
        if align:
            cell.alignment = align


def title_block(ws, text, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = TITLE
    c.fill = NAVY
    c.alignment = CENTER
    ws.row_dimensions[1].height = 28


# =========================================================
# 1) ورقة الافتراضات (INPUTS) - المرجع لكل الصيغ
# =========================================================
a = wb.active
a.title = "الافتراضات"
a.sheet_view.rightToLeft = True
title_block(a, "الافتراضات الأساسية (عدّل الخلايا الصفراء فقط)", 3)

a["A2"] = "البند"; a["B2"] = "القيمة"; a["C2"] = "الوحدة"
style_row(a, 2, 1, 3, fill=TEAL, font=HEAD, align=CENTER)

inputs = [
    ("سعر الصرف (USD/SAR)", 3.75, "ريال/دولار", "fx"),
    ("عدد أجهزة Reformer", 6, "جهاز", "machines"),
    ("سعة الجهاز (عميلة/حصة)", 1, "عميلة", "seat_per_machine"),
    ("عدد الحصص اليومية", 7, "حصة", "classes_day"),
    ("أيام التشغيل شهرياً", 26, "يوم", "days"),
    ("سعر الحصة المفردة", 150, "ريال", "price"),
    ("الإيجار السنوي", 204000, "ريال/سنة", "rent_year"),
    ("رواتب الموظفات (شهري)", 27000, "ريال/شهر", "salaries"),
    ("التأمينات + نهاية الخدمة (شهري)", 2200, "ريال/شهر", "gosi"),
    ("كهرباء وماء (شهري)", 3800, "ريال/شهر", "utilities"),
    ("نظام الحجز + اتصالات (شهري)", 1500, "ريال/شهر", "software"),
    ("التسويق (شهري)", 5000, "ريال/شهر", "marketing"),
    ("مستهلكات وغسيل (شهري)", 2000, "ريال/شهر", "supplies"),
    ("صيانة الأجهزة (شهري)", 1000, "ريال/شهر", "maint"),
    ("محاسبة وإداريات (شهري)", 1500, "ريال/شهر", "admin"),
    ("معدل الخصم لحساب NPV", 0.15, "نسبة", "discount"),
]
ref = {}
r = 3
for label, val, unit, key in inputs:
    a.cell(row=r, column=1, value=label).font = NORM
    vcell = a.cell(row=r, column=2, value=val)
    vcell.font = INPUT_FONT; vcell.fill = INPUT_FILL; vcell.alignment = CENTER
    a.cell(row=r, column=3, value=unit).font = NORM
    style_row(a, r, 1, 3, border=True)
    a.cell(row=r, column=2).fill = INPUT_FILL
    ref[key] = f"الافتراضات!$B${r}"
    r += 1

a.column_dimensions["A"].width = 34
a.column_dimensions["B"].width = 14
a.column_dimensions["C"].width = 14

# طاقة محسوبة
a.cell(row=r, column=1, value="الطاقة الشهرية القصوى (حصة-عميلة)").font = BOLD
a.cell(row=r, column=2,
       value=f"={ref['machines']}*{ref['seat_per_machine']}*{ref['classes_day']}*{ref['days']}").font = BOLD
style_row(a, r, 1, 3, fill=LGREY, font=BOLD); cap_ref = f"الافتراضات!$B${r}"; r += 1
a.cell(row=r, column=1, value="الإيراد الشهري عند 100% إشغال").font = BOLD
a.cell(row=r, column=2, value=f"={cap_ref}*{ref['price']}").font = BOLD
style_row(a, r, 1, 3, fill=LGREY, font=BOLD); maxrev_ref = f"الافتراضات!$B${r}"; r += 1

# =========================================================
# 2) التكلفة التأسيسية (CapEx)
# =========================================================
c = wb.create_sheet("التكلفة التأسيسية")
c.sheet_view.rightToLeft = True
title_block(c, "التكلفة التأسيسية (CapEx) — بالريال السعودي", 2)
c["A2"] = "البند"; c["B2"] = "القيمة (ريال)"
style_row(c, 2, 1, 2, fill=TEAL, font=HEAD, align=CENTER)
capex = [
    ("أجهزة Reformer مُورّدة (FOB+شحن+جمارك+نقل)", 41000),
    ("أدوات مساعدة (Props)", 15000),
    ("ديكور وتجهيز داخلي وأثاث", 80000),
    ("لوحة خارجية وأنظمة تقنية وكاميرات", 15000),
    ("رخص وتراخيص حكومية", 15000),
    ("احتياطي طوارئ", 9000),
]
r = 3
for label, val in capex:
    c.cell(row=r, column=1, value=label).font = NORM
    c.cell(row=r, column=2, value=val).font = NORM
    style_row(c, r, 1, 2); r += 1
c.cell(row=r, column=1, value="إجمالي التأسيس (قبل رأس المال العامل)").font = BOLD
c.cell(row=r, column=2, value=f"=SUM(B3:B{r-1})").font = BOLD
style_row(c, r, 1, 2, fill=LGREY, font=BOLD); r += 1
c.cell(row=r, column=1, value="رأس المال العامل").font = NORM
c.cell(row=r, column=2, value=100000).font = NORM
style_row(c, r, 1, 2); wc_row = r; r += 1
c.cell(row=r, column=1, value="إجمالي الاستثمار المطلوب").font = HEAD
c.cell(row=r, column=2, value=f"=B{r-2}+B{wc_row}").font = HEAD
style_row(c, r, 1, 2, fill=NAVY, font=HEAD)
invest_ref = f"'التكلفة التأسيسية'!$B${r}"
c.column_dimensions["A"].width = 42; c.column_dimensions["B"].width = 16

# =========================================================
# 3) التكاليف التشغيلية الشهرية (OpEx)
# =========================================================
o = wb.create_sheet("التكاليف التشغيلية")
o.sheet_view.rightToLeft = True
title_block(o, "التكاليف التشغيلية الشهرية (OpEx)", 2)
o["A2"] = "البند"; o["B2"] = "القيمة (ريال/شهر)"
style_row(o, 2, 1, 2, fill=TEAL, font=HEAD, align=CENTER)
opex_items = [
    ("الإيجار (سنوي ÷ 12)", f"={ref['rent_year']}/12"),
    ("رواتب الموظفات", f"={ref['salaries']}"),
    ("التأمينات + نهاية الخدمة", f"={ref['gosi']}"),
    ("كهرباء وماء", f"={ref['utilities']}"),
    ("نظام الحجز + اتصالات", f"={ref['software']}"),
    ("التسويق", f"={ref['marketing']}"),
    ("مستهلكات وغسيل وتنظيف", f"={ref['supplies']}"),
    ("صيانة الأجهزة", f"={ref['maint']}"),
    ("محاسبة وإداريات", f"={ref['admin']}"),
]
r = 3
for label, formula in opex_items:
    o.cell(row=r, column=1, value=label).font = NORM
    o.cell(row=r, column=2, value=formula).font = NORM
    style_row(o, r, 1, 2); r += 1
o.cell(row=r, column=1, value="إجمالي المصروف التشغيلي الشهري").font = HEAD
o.cell(row=r, column=2, value=f"=SUM(B3:B{r-1})").font = HEAD
style_row(o, r, 1, 2, fill=NAVY, font=HEAD)
opex_ref = f"'التكاليف التشغيلية'!$B${r}"
o.column_dimensions["A"].width = 34; o.column_dimensions["B"].width = 18

# =========================================================
# 4) سيناريوهات الإشغال
# =========================================================
s = wb.create_sheet("سيناريوهات الإشغال")
s.sheet_view.rightToLeft = True
title_block(s, "سيناريوهات الإشغال (الأرقام تتحدّث تلقائياً من الافتراضات)", 5)
headers = ["المؤشر", "إشغال 30%", "إشغال 50%", "إشغال 70%", "إشغال 90%"]
for i, h in enumerate(headers, start=1):
    cell = s.cell(row=2, column=i, value=h)
    cell.font = HEAD; cell.fill = TEAL; cell.alignment = CENTER; cell.border = BORDER
occ = {2: 0.30, 3: 0.50, 4: 0.70, 5: 0.90}
# صفوف
s.cell(row=3, column=1, value="نسبة الإشغال").font = BOLD
for col, v in occ.items():
    s.cell(row=3, column=col, value=v).font = NORM
s.cell(row=4, column=1, value="المقاعد المباعة/شهر").font = BOLD
for col in occ:
    L = get_column_letter(col)
    s.cell(row=4, column=col, value=f"=ROUND({L}3*{cap_ref},0)").font = NORM
s.cell(row=5, column=1, value="متوسط العميلات/حصة").font = BOLD
for col in occ:
    L = get_column_letter(col)
    s.cell(row=5, column=col, value=f"=ROUND({L}3*{ref['machines']},1)").font = NORM
s.cell(row=6, column=1, value="الإيراد الشهري (ريال)").font = BOLD
for col in occ:
    L = get_column_letter(col)
    s.cell(row=6, column=col, value=f"={L}4*{ref['price']}").font = NORM
s.cell(row=7, column=1, value="المصروف التشغيلي الشهري").font = BOLD
for col in occ:
    s.cell(row=7, column=col, value=f"=-{opex_ref}").font = NORM
s.cell(row=8, column=1, value="صافي الربح الشهري (ريال)").font = BOLD
for col in occ:
    L = get_column_letter(col)
    s.cell(row=8, column=col, value=f"={L}6+{L}7").font = BOLD
s.cell(row=9, column=1, value="صافي الربح السنوي (ريال)").font = BOLD
for col in occ:
    L = get_column_letter(col)
    s.cell(row=9, column=col, value=f"={L}8*12").font = BOLD
s.cell(row=10, column=1, value="فترة الاسترداد (شهر)").font = BOLD
for col in occ:
    L = get_column_letter(col)
    s.cell(row=10, column=col, value=f'=IF({L}8>0,ROUND({invest_ref}/{L}8,1),"لا يُسترد")').font = NORM
s.cell(row=11, column=1, value="ROI سنوي").font = BOLD
for col in occ:
    L = get_column_letter(col)
    s.cell(row=11, column=col, value=f"={L}9/{invest_ref}").font = NORM
for row in range(3, 12):
    style_row(s, row, 1, 5, align=CENTER)
    s.cell(row=row, column=1).alignment = RIGHT
# تنسيق ألوان صافي الربح
for col in occ:
    s.cell(row=8, column=col).fill = GREEN
    s.cell(row=9, column=col).fill = GREEN
# نِسب مئوية
for col in occ:
    s.cell(row=3, column=col).number_format = "0%"
    s.cell(row=11, column=col).number_format = "0%"
# نقطة التعادل
s.cell(row=13, column=1, value="نقطة التعادل (نسبة إشغال)").font = HEAD
s.cell(row=13, column=2, value=f"={opex_ref}/{maxrev_ref}").font = HEAD
s.cell(row=13, column=2).number_format = "0.0%"
style_row(s, 13, 1, 2, fill=LGREY, font=HEAD)
s.column_dimensions["A"].width = 26
for col in "BCDE":
    s.column_dimensions[col].width = 14

# =========================================================
# 5) التوقعات 5 سنوات + المؤشرات
# =========================================================
p = wb.create_sheet("توقعات 5 سنوات")
p.sheet_view.rightToLeft = True
title_block(p, "التوقعات المالية 5 سنوات والمؤشرات (NPV / IRR / Payback)", 6)
cols = ["البند", "السنة 0", "السنة 1", "السنة 2", "السنة 3", "السنة 4", "السنة 5"]
for i, h in enumerate(cols, start=1):
    cell = p.cell(row=2, column=i, value=h)
    cell.font = HEAD; cell.fill = TEAL; cell.alignment = CENTER; cell.border = BORDER
# الإشغال المتوقع
p.cell(row=3, column=1, value="متوسط الإشغال").font = BOLD
occ_ramp = [None, 0.38, 0.55, 0.63, 0.68, 0.70]
for yr in range(1, 6):
    p.cell(row=3, column=2 + yr, value=occ_ramp[yr]).font = INPUT_FONT
    p.cell(row=3, column=2 + yr).fill = INPUT_FILL
    p.cell(row=3, column=2 + yr).number_format = "0%"
# الإيراد السنوي = إشغال × طاقة × سعر × 12
p.cell(row=4, column=1, value="الإيراد السنوي").font = BOLD
for yr in range(1, 6):
    L = get_column_letter(2 + yr)
    p.cell(row=4, column=2 + yr, value=f"={L}3*{cap_ref}*{ref['price']}*12").font = NORM
# المصروف السنوي (إدخال قابل للتعديل، يعكس نمو الرواتب وإضافة مدربة)
p.cell(row=5, column=1, value="المصروف السنوي").font = BOLD
opex_year = [None, 762000, 754000, 885000, 911000, 938000]
for yr in range(1, 6):
    p.cell(row=5, column=2 + yr, value=opex_year[yr]).font = INPUT_FONT
    p.cell(row=5, column=2 + yr).fill = INPUT_FILL
# صافي التدفق النقدي
p.cell(row=6, column=1, value="صافي التدفق النقدي").font = BOLD
p.cell(row=6, column=2, value=f"=-{invest_ref}").font = BOLD  # السنة 0 = الاستثمار
p.cell(row=6, column=2).fill = RED
for yr in range(1, 6):
    L = get_column_letter(2 + yr)
    cell = p.cell(row=6, column=2 + yr, value=f"={L}4-{L}5")
    cell.font = BOLD; cell.fill = GREEN
# التدفق التراكمي
p.cell(row=7, column=1, value="التدفق التراكمي").font = BOLD
p.cell(row=7, column=2, value="=B6").font = NORM
for yr in range(1, 6):
    Lp = get_column_letter(1 + yr); L = get_column_letter(2 + yr)
    p.cell(row=7, column=2 + yr, value=f"={Lp}7+{L}6").font = NORM
for row in range(3, 8):
    style_row(p, row, 1, 7, align=CENTER)
    p.cell(row=row, column=1).alignment = RIGHT

# المؤشرات
p.cell(row=9, column=1, value="المؤشرات المالية").font = HEAD
p.cell(row=9, column=1).fill = NAVY
style_row(p, 9, 1, 2, fill=NAVY, font=HEAD)
metrics = [
    ("NPV (صافي القيمة الحالية)", f"=B6+NPV({ref['discount']},C6:G6)"),
    ("IRR (معدل العائد الداخلي)", "=IRR(B6:G6)"),
    ("إجمالي صافي التدفقات (5 سنوات)", "=SUM(C6:G6)"),
    ("نقطة التعادل (إشغال)", f"={opex_ref}*12/({cap_ref}*{ref['price']}*12)"),
]
r = 10
for label, formula in metrics:
    p.cell(row=r, column=1, value=label).font = BOLD
    p.cell(row=r, column=2, value=formula).font = BOLD
    style_row(p, r, 1, 2, fill=LGREY, font=BOLD)
    r += 1
p.cell(row=11, column=2).number_format = "0.0%"  # IRR
p.cell(row=13, column=2).number_format = "0.0%"  # break-even
p.column_dimensions["A"].width = 30
for col in "BCDEFG":
    p.column_dimensions[col].width = 13

# تنسيق الأرقام (فواصل آلاف) لكل الأوراق
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.number_format == "General":
                if abs(cell.value) >= 100:
                    cell.number_format = "#,##0"
            if isinstance(cell.value, str) and cell.value.startswith("=") \
               and cell.number_format == "General":
                cell.number_format = "#,##0"

# IRR/نسب: أعد ضبط ما يجب أن يكون نسبة (تم ضبطه أعلاه يدوياً)
import os
out = os.path.join(os.path.dirname(__file__), "النموذج-المالي-بيلاتس.xlsx")
wb.save(out)
print("تم إنشاء:", out)
